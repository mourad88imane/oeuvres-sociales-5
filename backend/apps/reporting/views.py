import csv
import io
import logging
from datetime import timedelta

from rest_framework import status, viewsets
from rest_framework.decorators import action
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

from core.pagination import StandardResultsSetPagination
from core.permissions import IsAdminOrGestionnaire, IsAdminOrReadOnly
from django.apps import apps
from django.db import models as db_models
from django.utils import timezone
from shared.ai.services import PredictionService

from .models import (
    DashboardWidget,
    DataExport,
    KpiDefinition,
    KpiSnapshot,
    ReportDefinition,
)
from .serializers import (
    DashboardWidgetSerializer,
    DataExportSerializer,
    KpiDefinitionSerializer,
    KpiSnapshotSerializer,
    ReportDefinitionSerializer,
    ReportGenerateSerializer,
)
from .services import aggregation_service

prediction_service = PredictionService()

logger = logging.getLogger("apps.reporting")


class AnalyticsViewSet(viewsets.ViewSet):
    permission_classes = [IsAuthenticated, IsAdminOrGestionnaire]

    @action(detail=False, methods=["get"])
    def summary(self, request):
        stats = aggregation_service.get_global_stats(user=request.user)
        return Response({"status": "success", "data": stats})

    @action(detail=False, methods=["get"])
    def trends(self, request):
        months = int(request.query_params.get("months", 12))
        trends = aggregation_service.get_monthly_trends(months=months)
        return Response({"status": "success", "data": trends})

    @action(detail=False, methods=["get"])
    def top(self, request):
        limit = int(request.query_params.get("limit", 10))
        top = aggregation_service.get_top_stats(limit=limit)
        return Response({"status": "success", "data": top})

    @action(detail=False, methods=["get"])
    def global_data(self, request):
        months = int(request.query_params.get("months", 12))
        limit = int(request.query_params.get("limit", 10))
        return Response(
            {
                "status": "success",
                "data": {
                    "summary": aggregation_service.get_global_stats(user=request.user),
                    "trends": aggregation_service.get_monthly_trends(months=months),
                    "top": aggregation_service.get_top_stats(limit=limit),
                    "kpis": aggregation_service.get_all_kpis(),
                },
            }
        )

    @action(detail=False, methods=["get"])
    def predictions(self, request):
        months_ahead = int(request.query_params.get("months", 3))
        return Response(
            {
                "status": "success",
                "data": {
                    "benefits_trend": prediction_service.predict_benefits_trend(
                        months_ahead=months_ahead
                    ),
                    "budget_consumption": prediction_service.predict_budget_consumption(),
                },
            }
        )

    @action(detail=False, methods=["get"])
    def anomalies(self, request):
        days = int(request.query_params.get("days", 90))
        return Response(
            {
                "status": "success",
                "data": prediction_service.detect_anomalies(days=days),
            }
        )

    @action(detail=False, methods=["get"])
    def decisions(self, request):
        return Response(
            {
                "status": "success",
                "data": prediction_service.decision_support(),
            }
        )


class KpiViewSet(viewsets.ModelViewSet):
    queryset = KpiDefinition.objects.filter(is_deleted=False)
    serializer_class = KpiDefinitionSerializer
    permission_classes = [IsAuthenticated, IsAdminOrReadOnly]

    @action(detail=False, methods=["get"])
    def values(self, request):
        kpis = aggregation_service.get_all_kpis()
        return Response({"status": "success", "data": kpis})

    @action(detail=False, methods=["post"])
    def snapshot(self, request):
        data = aggregation_service.compute_and_snapshot_kpis()
        return Response({"status": "success", "data": data})

    @action(detail=False, methods=["get"])
    def history(self, request):
        code = request.query_params.get("code", "")
        days = int(request.query_params.get("days", 90))
        if not code:
            return Response({"status": "error", "message": "Paramètre 'code' requis."}, status=400)
        kpi = KpiDefinition.objects.filter(code=code).first()
        if not kpi:
            return Response(
                {"status": "error", "message": f"KPI '{code}' introuvable."}, status=404
            )
        snapshots = KpiSnapshot.objects.filter(
            kpi=kpi,
            date__gte=timezone.localdate() - timedelta(days=days),
        ).order_by("date")
        serializer = KpiSnapshotSerializer(snapshots, many=True)
        return Response({"status": "success", "data": serializer.data})


class DashboardConfigViewSet(viewsets.ModelViewSet):
    queryset = DashboardWidget.objects.filter(is_deleted=False, is_active=True)
    serializer_class = DashboardWidgetSerializer
    permission_classes = [IsAuthenticated, IsAdminOrGestionnaire]

    def get_queryset(self):
        qs = super().get_queryset()
        return qs.filter(db_models.Q(is_global=True) | db_models.Q(user=self.request.user))

    def perform_create(self, serializer):
        serializer.save(user=self.request.user)

    @action(detail=False, methods=["get"])
    def my_widgets(self, request):
        widgets = (
            DashboardWidget.objects.filter(
                is_deleted=False,
                is_active=True,
            )
            .filter(db_models.Q(is_global=True) | db_models.Q(user=request.user))
            .order_by("display_order")
        )
        serializer = self.get_serializer(widgets, many=True)
        return Response({"status": "success", "data": serializer.data})


class ReportViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = ReportDefinition.objects.filter(is_deleted=False, is_active=True)
    serializer_class = ReportDefinitionSerializer
    pagination_class = StandardResultsSetPagination
    permission_classes = [IsAuthenticated, IsAdminOrGestionnaire]

    def get_queryset(self):
        qs = super().get_queryset()
        category = self.request.query_params.get("category", "")
        if category:
            qs = qs.filter(category=category)
        return qs

    @action(detail=True, methods=["post"])
    def generate(self, request, pk=None):
        report = self.get_object()
        ser = ReportGenerateSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        fmt = ser.validated_data.get("format", report.default_format)
        valid_formats = [f[0] for f in ReportDefinition.FORMATS]
        if fmt not in valid_formats:
            fmt = "csv"
        filters = ser.validated_data.get("filters", {})

        export = DataExport.objects.create(
            report=report,
            export_format=fmt,
            status="processing",
            filters_used=filters,
            created_by=request.user,
        )

        try:
            start = timezone.now()
            data = self._generate_report_data(report, filters)
            output = self._render_report(data, fmt, report)
            elapsed = (timezone.now() - start).total_seconds() * 1000

            ext = "xlsx" if fmt == "excel" else fmt
            filename = f"{report.code}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.{ext}"
            export.file.save(filename, output, save=True)
            export.status = "completed"
            export.row_count = len(data)
            export.file_size = export.file.size
            export.completed_at = timezone.now()
            export.duration_ms = int(elapsed)
            export.save(
                update_fields=[
                    "file",
                    "status",
                    "row_count",
                    "file_size",
                    "completed_at",
                    "duration_ms",
                    "updated_at",
                ]
            )
        except Exception as e:
            export.status = "failed"
            export.error_message = str(e)
            export.save(update_fields=["status", "error_message", "updated_at"])
            logger.exception("Échec génération rapport %s", report.code)
            return Response(
                {"status": "error", "message": str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

        out = DataExportSerializer(export, context={"request": request})
        return Response({"status": "success", "data": out.data})

    def _generate_report_data(self, report, filters):
        if report.category == "finance":
            return self._finance_data(filters)
        elif report.category == "employees":
            return self._employee_data(filters)
        elif report.category == "benefits":
            return self._benefit_data(filters)
        elif report.category == "conventions":
            return self._convention_data(filters)
        elif report.category == "kpi":
            return self._kpi_data()
        return []

    def _finance_data(self, filters):
        Payment = apps.get_model("finance", "Payment")
        qs = Payment.objects.alive().select_related("employee", "benefit", "budget", "fiscal_year")
        date_from = filters.get("date_from")
        date_to = filters.get("date_to")
        if date_from:
            qs = qs.filter(executed_date__gte=date_from)
        if date_to:
            qs = qs.filter(executed_date__lte=date_to)
        return list(
            qs.values(
                "reference",
                "employee__first_name",
                "employee__last_name",
                "employee__matricule",
                "benefit__reference",
                "amount",
                "status",
                "payment_method",
                "executed_date",
                "bank_reference",
            )[:5000]
        )

    def _employee_data(self, filters):
        Employee = apps.get_model("employees", "Employee")
        qs = Employee.objects.alive().select_related("department")
        return list(
            qs.values(
                "matricule",
                "first_name",
                "last_name",
                "email_professional",
                "phone",
                "department__name",
                "job_title",
                "wilaya",
                "status",
                "date_hired",
                "gender",
            )[:5000]
        )

    def _benefit_data(self, filters):
        Benefit = apps.get_model("benefits", "Benefit")
        qs = Benefit.objects.alive().select_related("employee", "benefit_type")
        date_from = filters.get("date_from")
        date_to = filters.get("date_to")
        if date_from:
            qs = qs.filter(created_at__date__gte=date_from)
        if date_to:
            qs = qs.filter(created_at__date__lte=date_to)
        return list(
            qs.values(
                "reference",
                "employee__first_name",
                "employee__last_name",
                "employee__matricule",
                "benefit_type__name",
                "requested_amount",
                "workflow_state",
                "created_at",
                "title",
            )[:5000]
        )

    def _convention_data(self, filters):
        Convention = apps.get_model("conventions", "Convention")
        qs = Convention.objects.alive().select_related("partner")
        return list(
            qs.values(
                "reference",
                "title",
                "partner__name",
                "partner__code",
                "status",
                "start_date",
                "end_date",
                "amount",
            )[:5000]
        )

    def _kpi_data(self):
        kpis = aggregation_service.get_all_kpis()
        return [
            {
                "code": k["code"],
                "name": k["name"],
                "value": k["current_value"],
                "previous": k["previous_value"],
                "variation": k["variation"],
                "target": k["target_value"],
                "trend": k["trend"],
            }
            for k in kpis
        ]

    def _render_report(self, data, fmt, report):
        if fmt == "csv":
            return self._render_csv(data, report)
        elif fmt == "json":
            return self._render_json(data)
        elif fmt == "pdf":
            return self._render_pdf(data, report)
        elif fmt in ("excel", "xlsx"):
            return self._render_excel(data, report)
        else:
            return self._render_csv(data, report)

    def _render_csv(self, data, report):
        if not data:
            output = io.BytesIO()
            output.write(b"Aucune donnee\n")
            output.seek(0)
            return output
        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=data[0].keys())
        writer.writeheader()
        writer.writerows(data)
        bytes_output = io.BytesIO(output.getvalue().encode("utf-8-sig"))
        return bytes_output

    def _render_json(self, data):
        import json

        output = io.BytesIO(json.dumps(data, indent=2, default=str).encode("utf-8"))
        return output

    def _render_excel(self, data, report):
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill

        wb = Workbook()
        ws = wb.active
        ws.title = report.code[:31] or "Report"

        if not data:
            ws.cell(row=1, column=1, value="Aucune donnee")
            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            return buf

        headers = list(data[0].keys())
        header_fill = PatternFill("solid", fgColor="1A3C6E")
        header_font = Font(bold=True, color="FFFFFF", size=10)

        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=ci, value=h.replace("__", " ").replace("_", " ").title())
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for ri, row in enumerate(data, 2):
            for ci, key in enumerate(headers, 1):
                val = row.get(key)
                cell = ws.cell(row=ri, column=ci, value=val if val is not None else "")
                cell.font = Font(size=9.5)
                if ri % 2 == 0:
                    cell.fill = PatternFill("solid", fgColor="F9FAFB")

        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 45)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    def _render_pdf(self, data, report):
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.lib.units import mm
        from reportlab.platypus import (
            PageBreak,
            Paragraph,
            SimpleDocTemplate,
            Spacer,
            Table,
            TableStyle,
        )

        buf = io.BytesIO()
        is_wide = len(data) > 20
        pagesize = landscape(A4) if is_wide else A4
        doc = SimpleDocTemplate(
            buf,
            pagesize=pagesize,
            topMargin=15 * mm,
            bottomMargin=20 * mm,
            leftMargin=15 * mm,
            rightMargin=15 * mm,
        )

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            "ReportTitle",
            parent=styles["Heading1"],
            fontSize=16,
            spaceAfter=6,
            spaceBefore=0,
        )
        subtitle_style = ParagraphStyle(
            "ReportSubtitle",
            parent=styles["Normal"],
            fontSize=9,
            textColor=colors.HexColor("#666666"),
            spaceAfter=12,
        )
        header_style = ParagraphStyle(
            "TableHeader",
            parent=styles["Normal"],
            fontSize=8,
            textColor=colors.white,
            alignment=1,
        )
        cell_style = ParagraphStyle(
            "TableCell",
            parent=styles["Normal"],
            fontSize=7.5,
            leading=10,
        )

        elements = []

        # Title
        elements.append(Paragraph(report.title, title_style))
        elements.append(Paragraph("Généré le — page $PAGE / $TOTAL", subtitle_style))
        elements.append(Spacer(1, 6 * mm))

        if not data:
            elements.append(Paragraph("Aucune donnée disponible.", styles["Normal"]))
        else:
            keys = list(data[0].keys())
            # Clean up header labels
            headers = [k.replace("__", " ").replace("_", " ").title() for k in keys]
            col_count = len(headers)

            max_col_width = max(20 * mm, (pagesize[0] - 40 * mm) / max(col_count, 1))
            col_widths = [min(max_col_width, 22 * mm)] * col_count

            table_data = [[Paragraph(h, header_style) for h in headers]]

            for row in data:
                table_data.append([Paragraph(str(row.get(k, "") or ""), cell_style) for k in keys])

            # Split into chunks if too large
            chunk_size = 40
            for start in range(0, len(table_data), chunk_size):
                if start > 0:
                    elements.append(PageBreak())
                    # Repeat header on each page
                    chunk = [table_data[0]] + table_data[start : start + chunk_size]
                else:
                    chunk = table_data[start : start + chunk_size]

                t = Table(chunk, colWidths=col_widths, repeatRows=1)
                t.setStyle(
                    TableStyle(
                        [
                            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1A3C6E")),
                            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                            ("FONTSIZE", (0, 0), (-1, -1), 7.5),
                            ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                            ("ALIGN", (-1, 0), (-1, -1), "RIGHT"),
                            ("VALIGN", (0, 0), (-1, -1), "TOP"),
                            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E5E7EB")),
                            (
                                "ROWBACKGROUNDS",
                                (0, 1),
                                (-1, -1),
                                [colors.white, colors.HexColor("#F9FAFB")],
                            ),
                            ("TOPPADDING", (0, 0), (-1, -1), 4),
                            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                            ("LEFTPADDING", (0, 0), (-1, -1), 6),
                            ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                        ]
                    )
                )
                elements.append(t)

        doc.build(elements)
        buf.seek(0)
        return buf


class ExportViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = DataExport.objects.filter(is_deleted=False)
    serializer_class = DataExportSerializer
    pagination_class = StandardResultsSetPagination
    permission_classes = [IsAuthenticated, IsAdminOrGestionnaire]

    def get_queryset(self):
        qs = super().get_queryset()
        if not self.request.user.is_staff:
            qs = qs.filter(created_by=self.request.user)
        return qs
