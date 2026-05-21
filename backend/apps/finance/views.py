"""Finance views — API REST complète."""

import io
import logging

from rest_framework.views import APIView

from rest_framework import status
from rest_framework.decorators import action
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.viewsets import ModelViewSet, ReadOnlyModelViewSet

from core.pagination import StandardResultsSetPagination
from core.permissions import IsAdmin, IsAdminOrComptable, IsAdminOrGestionnaire
from django.http import HttpResponse
from django.utils import timezone

from .models import Budget, FinancialAlert, FinancialEntry, FiscalYear, PaymentBatch
from .serializers import (
    BatchCreateSerializer,
    BudgetCreateSerializer,
    BudgetListSerializer,
    BudgetSerializer,
    FinancialAlertSerializer,
    FinancialEntrySerializer,
    FiscalYearCreateSerializer,
    FiscalYearListSerializer,
    FiscalYearSerializer,
    PaymentBatchSerializer,
    PaymentCancelSerializer,
    PaymentCreateSerializer,
    PaymentListSerializer,
    PaymentMarkPaidSerializer,
    PaymentSerializer,
)
from .services import (
    BudgetService,
    FinanceValidationError,
    FinancialReportService,
    FiscalYearService,
    PaymentBatchService,
    PaymentService,
)

logger = logging.getLogger("apps.finance")


def finance_error_response(e: FinanceValidationError):
    return Response(
        {"status": "error", "code": getattr(e, "code", "FINANCE_ERROR"), "message": e.message},
        status=status.HTTP_400_BAD_REQUEST,
    )


# ═══════════════════════════════════════════════════════════
# FISCAL YEAR
# ═══════════════════════════════════════════════════════════
class FiscalYearViewSet(ModelViewSet):
    queryset = FiscalYear.objects.filter(is_deleted=False)
    pagination_class = StandardResultsSetPagination

    def get_serializer_class(self):
        if self.action == "list":
            return FiscalYearListSerializer
        if self.action == "create":
            return FiscalYearCreateSerializer
        return FiscalYearSerializer

    def get_permissions(self):
        if self.action in ("list", "retrieve", "dashboard", "active"):
            return [IsAuthenticated()]
        return [IsAuthenticated(), IsAdmin()]

    svc = FiscalYearService()

    def create(self, request, *args, **kwargs):
        ser = FiscalYearCreateSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        try:
            fy = self.svc.create(ser.validated_data, user=request.user, request=request)
        except FinanceValidationError as e:
            return finance_error_response(e)
        return Response(
            {"status": "success", "data": FiscalYearSerializer(fy).data},
            status=status.HTTP_201_CREATED,
        )

    @action(detail=False, methods=["get"], url_path="active")
    def active(self, request):
        fy = self.svc.get_active_year()
        if not fy:
            return Response(
                {"status": "success", "data": None, "message": "Aucun exercice fiscal actif."}
            )
        return Response({"status": "success", "data": FiscalYearSerializer(fy).data})

    @action(detail=True, methods=["post"], url_path="open")
    def open_year(self, request, pk=None):
        fy = self.get_object()
        try:
            fy = self.svc.open(fy, user=request.user)
        except FinanceValidationError as e:
            return finance_error_response(e)
        return Response({"status": "success", "data": FiscalYearSerializer(fy).data})

    @action(detail=True, methods=["post"], url_path="close")
    def close_year(self, request, pk=None):
        fy = self.get_object()
        try:
            fy = self.svc.close(fy, user=request.user)
        except FinanceValidationError as e:
            return finance_error_response(e)
        return Response({"status": "success", "data": FiscalYearSerializer(fy).data})

    @action(detail=True, methods=["get"], url_path="dashboard")
    def dashboard(self, request, pk=None):
        fy = self.get_object()
        data = FinancialReportService().get_dashboard_summary(str(fy.id))
        return Response({"status": "success", "data": data})


# ═══════════════════════════════════════════════════════════
# BUDGET
# ═══════════════════════════════════════════════════════════
class BudgetViewSet(ModelViewSet):
    pagination_class = StandardResultsSetPagination
    svc = BudgetService()

    def get_serializer_class(self):
        if self.action == "list":
            return BudgetListSerializer
        if self.action == "create":
            return BudgetCreateSerializer
        return BudgetSerializer

    def get_permissions(self):
        if self.action in ("list", "retrieve", "consumption_report"):
            return [IsAuthenticated()]
        if self.action == "approve":
            return [IsAuthenticated(), IsAdmin()]
        return [IsAuthenticated(), IsAdminOrComptable()]

    def get_queryset(self):
        qs = self.svc.get_queryset()
        p = self.request.query_params
        if fy := p.get("fiscal_year"):
            qs = qs.filter(fiscal_year_id=fy)
        if bt := p.get("benefit_type"):
            qs = qs.filter(benefit_type_id=bt)
        if dept := p.get("department"):
            qs = qs.filter(department_id=dept)
        if st := p.get("status"):
            qs = qs.filter(status=st)
        return qs

    def create(self, request, *args, **kwargs):
        ser = BudgetCreateSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        try:
            b = self.svc.create(ser.validated_data, user=request.user, request=request)
        except FinanceValidationError as e:
            return finance_error_response(e)
        return Response(
            {"status": "success", "data": BudgetSerializer(b).data}, status=status.HTTP_201_CREATED
        )

    def update(self, request, *args, **kwargs):
        partial = kwargs.pop("partial", False)
        b = self.get_object()
        ser = BudgetCreateSerializer(b, data=request.data, partial=partial)
        ser.is_valid(raise_exception=True)
        try:
            b = self.svc.update(b, ser.validated_data, user=request.user, request=request)
        except FinanceValidationError as e:
            return finance_error_response(e)
        return Response({"status": "success", "data": BudgetSerializer(b).data})

    @action(detail=True, methods=["post"], url_path="approve")
    def approve(self, request, pk=None):
        b = self.get_object()
        try:
            b = self.svc.approve(b, user=request.user)
        except FinanceValidationError as e:
            return finance_error_response(e)
        return Response({"status": "success", "data": BudgetSerializer(b).data})

    @action(detail=False, methods=["get"], url_path="consumption-report")
    def consumption_report(self, request):
        fy_id = request.query_params.get("fiscal_year")
        if not fy_id:
            return Response({"status": "error", "message": "fiscal_year requis."}, status=400)
        data = self.svc.get_consumption_report(fy_id)
        return Response({"status": "success", "count": len(data), "data": data})

    @action(detail=False, methods=["get"], url_path="export")
    def export(self, request):
        fy_id = request.query_params.get("fiscal_year")
        if not fy_id:
            return Response({"status": "error", "message": "fiscal_year requis."}, status=400)
        from .services import FinancialReportService

        rows = FinancialReportService().get_budget_export_data(fy_id)
        return self._to_excel(rows, "budgets")

    def _to_excel(self, rows: list, name: str) -> HttpResponse:
        try:
            import openpyxl
            from openpyxl.styles import Alignment, Font, PatternFill

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = name.capitalize()
            if not rows:
                resp = HttpResponse(
                    content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
                resp["Content-Disposition"] = f'attachment; filename="{name}.xlsx"'
                wb.save(resp)
                return resp
            # En-têtes
            headers = list(rows[0].keys())
            for ci, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=ci, value=h)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="1A3C6E")
                cell.alignment = Alignment(horizontal="center")
            # Données
            for ri, row in enumerate(rows, 2):
                for ci, key in enumerate(headers, 1):
                    ws.cell(row=ri, column=ci, value=row[key])
            # Auto-width
            for col in ws.columns:
                max_len = max((len(str(c.value or "")) for c in col), default=10)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)
            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            resp = HttpResponse(
                buf.read(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            resp["Content-Disposition"] = f'attachment; filename="{name}.xlsx"'
            return resp
        except ImportError:
            # Fallback CSV si openpyxl absent
            import csv

            output = io.StringIO()
            w = csv.DictWriter(output, fieldnames=list(rows[0].keys()))
            w.writeheader()
            w.writerows(rows)
            resp = HttpResponse(output.getvalue(), content_type="text/csv; charset=utf-8-sig")
            resp["Content-Disposition"] = f'attachment; filename="{name}.csv"'
            return resp


# ═══════════════════════════════════════════════════════════
# PAYMENT
# ═══════════════════════════════════════════════════════════
class PaymentViewSet(ModelViewSet):
    pagination_class = StandardResultsSetPagination
    svc = PaymentService()

    def get_serializer_class(self):
        if self.action == "list":
            return PaymentListSerializer
        if self.action == "create":
            return PaymentCreateSerializer
        if self.action == "mark_paid":
            return PaymentMarkPaidSerializer
        if self.action == "cancel":
            return PaymentCancelSerializer
        return PaymentSerializer

    def get_permissions(self):
        if self.action in ("list", "retrieve", "statistics"):
            return [IsAuthenticated(), IsAdminOrGestionnaire()]
        if self.action in ("mark_paid", "approve", "batch_create"):
            return [IsAuthenticated(), IsAdminOrComptable()]
        return [IsAuthenticated(), IsAdminOrComptable()]

    def get_queryset(self):
        p = self.request.query_params
        return self.svc.search(
            queryset=self.svc.get_queryset(),
            search=p.get("search", ""),
            status=p.get("status", ""),
            fiscal_year_id=p.get("fiscal_year", ""),
            budget_id=p.get("budget", ""),
            employee_id=p.get("employee", ""),
            department_id=p.get("department", ""),
            date_from=p.get("date_from", ""),
            date_to=p.get("date_to", ""),
            anomaly_only=p.get("anomaly_only", "").lower() == "true",
            ordering=p.get("ordering", "-created_at"),
        )

    def create(self, request, *args, **kwargs):
        ser = PaymentCreateSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        d = ser.validated_data
        try:
            from apps.benefits.models import Benefit

            benefit = Benefit.objects.get(pk=d["benefit_id"])
            budget = Budget.objects.get(pk=d["budget_id"]) if d.get("budget_id") else None
            fy = FiscalYear.objects.get(pk=d["fiscal_year_id"]) if d.get("fiscal_year_id") else None
            pay = self.svc.create_from_benefit(
                benefit, budget, fy, user=request.user, request=request
            )
        except FinanceValidationError as e:
            return finance_error_response(e)
        except Exception as e:
            return Response({"status": "error", "message": str(e)}, status=400)
        return Response(
            {"status": "success", "data": PaymentSerializer(pay).data},
            status=status.HTTP_201_CREATED,
        )

    def retrieve(self, request, *args, **kwargs):
        p = self.get_object()
        return Response({"status": "success", "data": PaymentSerializer(p).data})

    @action(detail=True, methods=["post"], url_path="approve")
    def approve(self, request, pk=None):
        p = self.get_object()
        try:
            p = self.svc.approve(p, user=request.user, request=request)
        except FinanceValidationError as e:
            return finance_error_response(e)
        return Response({"status": "success", "data": PaymentSerializer(p).data})

    @action(detail=True, methods=["post"], url_path="mark-paid")
    def mark_paid(self, request, pk=None):
        p = self.get_object()
        ser = PaymentMarkPaidSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        d = ser.validated_data
        try:
            p = self.svc.mark_paid(
                p,
                d["bank_reference"],
                executed_date=d.get("executed_date"),
                paid_amount=d.get("paid_amount"),
                user=request.user,
                request=request,
            )
        except FinanceValidationError as e:
            return finance_error_response(e)
        return Response({"status": "success", "data": PaymentSerializer(p).data})

    @action(detail=True, methods=["post"], url_path="cancel")
    def cancel(self, request, pk=None):
        p = self.get_object()
        ser = PaymentCancelSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        try:
            p = self.svc.cancel(p, ser.validated_data["reason"], user=request.user, request=request)
        except FinanceValidationError as e:
            return finance_error_response(e)
        return Response({"status": "success", "data": PaymentSerializer(p).data})

    @action(detail=False, methods=["get"], url_path="statistics")
    def statistics(self, request):
        fy_id = request.query_params.get("fiscal_year")
        data = FinancialReportService().get_dashboard_summary(fy_id)
        return Response({"status": "success", "data": data})

    @action(detail=False, methods=["get"], url_path="export")
    def export(self, request):
        from shared.audit.services import AuditService

        p = request.query_params
        filters = {
            "search": p.get("search", ""),
            "status": p.get("status", ""),
            "fiscal_year_id": p.get("fiscal_year", ""),
            "budget_id": p.get("budget", ""),
            "date_from": p.get("date_from", ""),
            "date_to": p.get("date_to", ""),
        }
        AuditService().log_export(
            user=request.user, content_type="payments", filters=filters, request=request
        )
        rows = FinancialReportService().get_payment_export_data(
            {k: v for k, v in filters.items() if v}
        )
        return BudgetViewSet._to_excel(self, rows, "paiements")


# ═══════════════════════════════════════════════════════════
# PAYMENT BATCH
# ═══════════════════════════════════════════════════════════
class PaymentBatchViewSet(ModelViewSet):
    queryset = PaymentBatch.objects.filter(is_deleted=False).select_related(
        "fiscal_year", "approved_by"
    )
    serializer_class = PaymentBatchSerializer
    pagination_class = StandardResultsSetPagination
    svc = PaymentBatchService()

    def get_permissions(self):
        return [IsAuthenticated(), IsAdminOrComptable()]

    def create(self, request, *args, **kwargs):
        ser = BatchCreateSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        d = ser.validated_data
        try:
            fy = FiscalYear.objects.get(pk=d["fiscal_year_id"])
            batch = self.svc.create(
                data={
                    "label": d["label"],
                    "fiscal_year": fy,
                    "payment_method": d["payment_method"],
                    "scheduled_date": d.get("scheduled_date"),
                    "notes": d.get("notes", ""),
                },
                payment_ids=[str(pid) for pid in d["payment_ids"]],
                user=request.user,
                request=request,
            )
        except FinanceValidationError as e:
            return finance_error_response(e)
        return Response(
            {"status": "success", "data": PaymentBatchSerializer(batch).data},
            status=status.HTTP_201_CREATED,
        )

    @action(detail=True, methods=["post"], url_path="approve")
    def approve(self, request, pk=None):
        batch = self.get_object()
        try:
            batch = self.svc.approve(batch, user=request.user)
        except FinanceValidationError as e:
            return finance_error_response(e)
        return Response({"status": "success", "data": PaymentBatchSerializer(batch).data})


# ═══════════════════════════════════════════════════════════
# FINANCIAL ENTRY (lecture seule — journal comptable)
# ═══════════════════════════════════════════════════════════
class FinancialEntryViewSet(ReadOnlyModelViewSet):
    serializer_class = FinancialEntrySerializer
    pagination_class = StandardResultsSetPagination
    permission_classes = [IsAuthenticated, IsAdminOrComptable]

    def get_queryset(self):
        qs = FinancialEntry.objects.select_related(
            "fiscal_year", "payment", "budget", "created_by_user"
        )
        p = self.request.query_params
        if fy := p.get("fiscal_year"):
            qs = qs.filter(fiscal_year_id=fy)
        if et := p.get("entry_type"):
            qs = qs.filter(entry_type=et)
        if pay := p.get("payment"):
            qs = qs.filter(payment_id=pay)
        return qs.order_by("-entry_date", "-entry_number")


# ═══════════════════════════════════════════════════════════
# FINANCIAL ALERTS
# ═══════════════════════════════════════════════════════════
class FinancialAlertViewSet(ModelViewSet):
    serializer_class = FinancialAlertSerializer
    pagination_class = StandardResultsSetPagination

    def get_permissions(self):
        return [IsAuthenticated(), IsAdminOrComptable()]

    def get_queryset(self):
        qs = FinancialAlert.objects.filter(is_deleted=False)
        p = self.request.query_params
        if p.get("unresolved"):
            qs = qs.filter(is_resolved=False)
        if sv := p.get("severity"):
            qs = qs.filter(severity=sv)
        if fy := p.get("fiscal_year"):
            qs = qs.filter(fiscal_year_id=fy)
        return qs.order_by("-created_at")

    @action(detail=True, methods=["post"], url_path="resolve")
    def resolve(self, request, pk=None):
        alert = self.get_object()
        alert.is_resolved = True
        alert.resolved_at = timezone.now()
        alert.resolved_by = request.user
        alert.save(update_fields=["is_resolved", "resolved_at", "resolved_by", "updated_at"])
        return Response({"status": "success", "message": "Alerte résolue."})

    @action(detail=False, methods=["post"], url_path="mark-all-read")
    def mark_all_read(self, request):
        FinancialAlert.objects.filter(is_deleted=False, is_read=False).update(is_read=True)
        return Response({"status": "success", "message": "Toutes les alertes marquées comme lues."})


# ═══════════════════════════════════════════════════════════
# DASHBOARD GLOBAL
# ═══════════════════════════════════════════════════════════
class FinanceDashboardView(APIView):
    """Vue unique pour le dashboard finance — agrège toutes les données."""

    permission_classes = [IsAuthenticated, IsAdminOrComptable]

    def get(self, request):
        fy_id = request.query_params.get("fiscal_year")
        data = FinancialReportService().get_dashboard_summary(fy_id)
        return Response({"status": "success", "data": data})
